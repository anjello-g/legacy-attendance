import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from difflib import SequenceMatcher
from datetime import datetime, time, date
import dateutil.parser
import gc

st.set_page_config(page_title="Attendance Builder", page_icon="🗂️", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=Space+Grotesk:wght@400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
h1, h2, h3 { font-family: 'Space Grotesk', sans-serif; }
.stApp { background: #0f1117; color: #e8eaf0; }
.block-container { padding-top: 2rem; max-width: 1400px; }
.metric-card {
    background: #1a1d27; border: 1px solid #2a2d3a;
    border-radius: 12px; padding: 1.2rem 1.5rem; text-align: center;
}
.metric-val { font-family: 'Space Grotesk', sans-serif; font-size: 2rem; font-weight: 700; line-height: 1; }
.metric-lbl { font-size: 0.78rem; color: #7c7f8e; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.08em; }
.matched { color: #4ade80; } .unmatched { color: #f87171; } .fuzzy { color: #facc15; }
.neutral { color: #a78bfa; } .review { color: #fb923c; } .not-hired { color: #60a5fa; }
.step-badge {
    display: inline-block; background: #1e1b4b; border: 1px solid #4338ca;
    color: #a5b4fc; border-radius: 8px; padding: 0.3rem 0.8rem; font-size: 0.72rem;
    font-family: 'Space Grotesk', sans-serif; font-weight: 600; letter-spacing: 0.08em;
    text-transform: uppercase; margin-bottom: 0.8rem;
}
.step-badge.done { background: #14532d; border-color: #166534; color: #4ade80; }
div[data-testid="stFileUploader"] { background: #1a1d27; border: 2px dashed #2a2d3a; border-radius: 12px; padding: 1rem; }
div[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }
.stButton > button {
    background: linear-gradient(135deg, #6366f1, #8b5cf6); color: white; border: none;
    border-radius: 8px; padding: 0.6rem 2rem; font-family: 'Space Grotesk', sans-serif;
    font-weight: 600; font-size: 0.95rem; transition: opacity 0.2s; width: 100%;
}
.stButton > button:hover { opacity: 0.85; }
.section-header {
    font-family: 'Space Grotesk', sans-serif; font-size: 0.7rem; font-weight: 600;
    letter-spacing: 0.12em; text-transform: uppercase; color: #6366f1; margin-bottom: 0.5rem;
}
</style>
""", unsafe_allow_html=True)

for key in ["roster", "attendance", "merged_headcount"]:
    if key not in st.session_state:
        st.session_state[key] = None

EXCLUDED_LOCATIONS = {"clark", "dsi", "zamboanga", "isabela"}
DAY_COLS = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
FUZZY_THRESHOLD = 0.82
PREVIEW_ROWS = 50

# ═══════════════════════════════════════════════════════════════════════════════
# FAST HELPERS — all dict/hashmap based, no pandas iteration
# ═══════════════════════════════════════════════════════════════════════════════

def normalize(name: str) -> str:
    if pd.isna(name): return ""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z\s]", "", str(name).lower().strip()))

def is_active_hire_status(status) -> bool:
    """Return True if HireStatus indicates an active/current employee."""
    if pd.isna(status):
        return True  # Unknown = assume active
    s = str(status).lower().strip()
    # Active/rehired indicators
    active_terms = {"active", "rehired", "hired", "new hire", "onboarding", "probation", "regular", "permanent"}
    # Inactive/terminated indicators
    inactive_terms = {"terminated", "separated", "resigned", "inactive", "retired", "end of contract", "eoc", "awol", "blacklisted"}
    if any(t in s for t in active_terms):
        return True
    if any(t in s for t in inactive_terms):
        return False
    return True  # Default to active for unknown statuses

def build_staff_lookup(staff_df: pd.DataFrame) -> dict:
    """Build lookup that stores (index, hire_status_is_active) tuples keyed by EmployeeNumber."""
    lookup: dict[int, list[tuple[int, bool]]] = {}
    for idx, row in staff_df.iterrows():
        is_active = is_active_hire_status(row.get("HireStatus"))
        # Primary key: EmployeeNumber (numeric)
        emp_num = row.get("EmployeeNumber")
        if pd.notna(emp_num):
            try:
                emp_int = int(float(emp_num))
                lookup.setdefault(emp_int, []).append((idx, is_active))
            except (ValueError, TypeError):
                pass
        # Fallback: normalized name
        key = normalize(row.get("Name", ""))
        if key:
            lookup.setdefault(key, []).append((idx, is_active))
        first = normalize(row.get("FirstName", ""))
        last = normalize(row.get("LastName", ""))
        if first and last:
            lookup.setdefault(f"{first} {last}", []).append((idx, is_active))
    return lookup

def match_employee(sched_row: pd.Series, lookup: dict, staff_df: pd.DataFrame):
    """Match schedule row to staff. Priority: EmployeeNumber exact > Name exact > Name fuzzy."""
    sched_emp_id = sched_row.get("EmployeeID")
    sched_name = sched_row.get("Name", "")
    sched_hire_status = sched_row.get("HireStatus", None)

    # Try 1: EmployeeID exact match (highest priority)
    if pd.notna(sched_emp_id):
        try:
            emp_int = int(float(sched_emp_id))
            if emp_int in lookup:
                matches = lookup[emp_int]
                if len(matches) == 1:
                    return staff_df.loc[matches[0][0]], "exact_id"
                # Multiple matches — prefer active over inactive
                active_matches = [m for m in matches if m[1]]
                if active_matches:
                    if sched_hire_status is not None:
                        sched_active = is_active_hire_status(sched_hire_status)
                        hinted = [m for m in active_matches if m[1] == sched_active]
                        if hinted:
                            return staff_df.loc[hinted[0][0]], "exact_id"
                    return staff_df.loc[active_matches[0][0]], "exact_id"
                return staff_df.loc[matches[0][0]], "exact_id"
        except (ValueError, TypeError):
            pass

    # Try 2: Name exact match (fallback)
    norm = normalize(sched_name)
    if not norm: return None, "no_name"

    if norm in lookup:
        matches = lookup[norm]
        # Filter out EmployeeNumber-keyed entries for name matching
        name_matches = [m for m in matches if isinstance(m, tuple) and len(m) == 2]
        if not name_matches:
            name_matches = matches
        if len(name_matches) == 1:
            return staff_df.loc[name_matches[0][0]], "exact_name"
        active_matches = [m for m in name_matches if m[1]]
        if active_matches:
            if sched_hire_status is not None:
                sched_active = is_active_hire_status(sched_hire_status)
                hinted = [m for m in active_matches if m[1] == sched_active]
                if hinted:
                    return staff_df.loc[hinted[0][0]], "exact_name"
            return staff_df.loc[active_matches[0][0]], "exact_name"
        return staff_df.loc[name_matches[0][0]], "exact_name"

    # Try 3: Fuzzy fallback
    best_score, best_key = 0.0, None
    for key in lookup:
        if isinstance(key, str):  # Only fuzzy match against string keys (names)
            sc = SequenceMatcher(None, norm, key).ratio()
            if sc > best_score:
                best_score, best_key = sc, key
    if best_score >= FUZZY_THRESHOLD and best_key:
        matches = lookup[best_key]
        name_matches = [m for m in matches if isinstance(m, tuple) and len(m) == 2]
        if not name_matches:
            name_matches = matches
        active_matches = [m for m in name_matches if m[1]]
        if active_matches:
            if sched_hire_status is not None:
                sched_active = is_active_hire_status(sched_hire_status)
                hinted = [m for m in active_matches if m[1] == sched_active]
                if hinted:
                    return staff_df.loc[hinted[0][0]], f"fuzzy({best_score:.0%})"
            return staff_df.loc[active_matches[0][0]], f"fuzzy({best_score:.0%})"
        return staff_df.loc[name_matches[0][0]], f"fuzzy({best_score:.0%})"
    return None, "unmatched"

def normalize_schedule(val) -> str:
    if pd.isna(val): return "Rest Day"
    s = str(val).strip()
    return "Rest Day" if s in ("", "0000 - 0000") else s

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
        df.to_excel(w, index=False, sheet_name="Result")
    return buf.getvalue()

def read_timesheet(file) -> pd.DataFrame:
    raw = pd.read_excel(file, sheet_name=0, header=None)
    header_row = next((i for i, row in raw.iterrows() if row.astype(str).str.contains("Agent Email", case=False).any()), None)
    if header_row is None:
        raise ValueError("Could not find 'Agent Email' header in timesheet.")
    df = pd.read_excel(file, sheet_name=0, header=header_row)
    keep = ["Agent", "Agent Email", "Active", "Break", "First Login", "Last Logout"]
    df = df[[c for c in keep if c in df.columns]].copy()
    df = df[df["Agent Email"].notna() & (df["Agent Email"].astype(str).str.strip() != "")]
    return df.reset_index(drop=True)

def read_leave_file(file) -> pd.DataFrame:
    df = pd.read_excel(file, sheet_name=0, header=0)
    if len(df) > 0:
        first_row_vals = [str(x).lower().strip() for x in df.iloc[0].values]
        if "name" in first_row_vals and "position" in first_row_vals:
            df = pd.read_excel(file, sheet_name=0, header=1)
    return df

def parse_date_str(date_str: str):
    for fmt in ("%m-%d-%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

def get_day_column(date_str: str) -> str:
    dt = parse_date_str(date_str)
    return dt.strftime("%a") if dt else None

def parse_doj_knack(value) -> date:
    if pd.isna(value) or str(value).strip() in ("", "NaT"):
        return None
    dt = safe_parse_datetime(value)
    return dt.date() if dt else None

def safe_parse_datetime(value):
    if pd.isna(value) or str(value).strip() in ("", "NaT"):
        return None
    if isinstance(value, datetime):
        return value
    try:
        return dateutil.parser.parse(str(value))
    except:
        return None

def parse_shift_start(shift_str: str) -> time:
    if pd.isna(shift_str) or shift_str == "Rest Day":
        return None
    parts = shift_str.strip().split("-")
    if len(parts) < 2:
        return None
    start_str = parts[0].strip()
    if len(start_str) == 4 and start_str.isdigit():
        h, m = int(start_str[:2]), int(start_str[2:])
        if 0 <= h <= 23 and 0 <= m <= 59:
            return time(h, m)
    return None

def compute_late(shift_start: time, first_login_dt, date_obj: date) -> tuple[int, float]:
    if shift_start is None or first_login_dt is None or date_obj is None:
        return 0, 0.0
    diff = (first_login_dt - datetime.combine(date_obj, shift_start)).total_seconds() / 60.0
    return (1, round(diff, 2)) if diff > 0 else (0, 0.0)

def normalize_filename_date(filename: str) -> str:
    base = filename.replace(".xlsx", "").replace(".xls", "").strip()
    dt = parse_date_str(base)
    return dt.strftime("%Y-%m-%d") if dt else base

def extract_headcount_date(filename: str, hc_df: pd.DataFrame) -> date:
    for col in ["Date Exported", "DateExported", "Export Date", "ExportDate", "Date", "As Of"]:
        if col in hc_df.columns:
            val = hc_df[col].dropna().iloc[0] if len(hc_df) > 0 else None
            if val is not None:
                dt = safe_parse_datetime(val)
                if dt:
                    return dt.date() if hasattr(dt, 'date') else dt
    dt = parse_date_str(filename.replace(".xlsx", "").replace(".xls", "").strip())
    return dt.date() if dt else None

def parse_override_date(value) -> date:
    if pd.isna(value) or str(value).strip() in ("", "NaT", "None"):
        return None
    dt = safe_parse_datetime(value)
    if dt:
        return dt.date() if hasattr(dt, 'date') else dt
    for fmt in ("%m-%d-%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y",
                "%m-%d-%y", "%d-%m-%y", "%y-%m-%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None

def clean_for_display(df: pd.DataFrame) -> pd.DataFrame:
    df = df.head(PREVIEW_ROWS).copy()
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).replace('nan', '').replace('None', '')
        elif pd.api.types.is_integer_dtype(df[col]) and df[col].isna().any():
            df[col] = df[col].astype('Int64').astype(str).replace('<NA>', '')
    return df

def build_override_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    buf = BytesIO()
    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4338ca", end_color="4338ca", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    def style_row(ws, row_idx, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = tb
    ws1 = wb.active; ws1.title = "Schedule"
    ws1.append(["ECN", "Effective Date", "Effective Until", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"])
    style_row(ws1, 1, 10)
    ws1.append([12345, "05-01-2026", "05-31-2026", "0900 - 1800", "Rest Day", "0900 - 1800", "0900 - 1800", "0900 - 1800", "0900 - 1800", "Rest Day"])
    for c in ['A','B','C']: ws1.column_dimensions[c].width = 16
    for c in ['D','E','F','G','H','I','J']: ws1.column_dimensions[c].width = 14
    ws2 = wb.create_sheet(title="Leave")
    ws2.append(["ECN", "Date", "Leave"]); style_row(ws2, 1, 3)
    ws2.append([12345, "05-20-2026", 1])
    for c in ['A','B','C']: ws2.column_dimensions[c].width = 14
    ws3 = wb.create_sheet(title="Holidays")
    ws3.append(["Project", "Sub-Process", "Date"]); style_row(ws3, 1, 3)
    ws3.append(["Project Alpha", "", "05-01-2026"])
    for c in ['A','B','C']: ws3.column_dimensions[c].width = 18
    ws4 = wb.create_sheet(title="Exemptions")
    ws4.append(["ECN", "Exemption Date", "Effective Until", "Active/Inactive"]); style_row(ws4, 1, 4)
    ws4.append([12345, "05-10-2026", "05-20-2026", "Inactive"])
    for c in ['A','B','C','D']: ws4.column_dimensions[c].width = 18
    wb.save(buf)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Roster
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("## 🗂️ Attendance Builder")
st.markdown('<p style="color:#7c7f8e;margin-top:-0.5rem;">Step 1: build the roster → Step 2: process everything in one go</p>', unsafe_allow_html=True)

tab1, tab2 = st.tabs(["📋 Step 1 — Schedule + Staff", "⏱️ Step 2 — Build Final Attendance"])

with tab1:
    st.markdown("")
    roster_ready = st.session_state.roster is not None
    badge_html = '<span class="step-badge done">✔ Roster ready</span>' if roster_ready else '<span class="step-badge">Roster not built yet</span>'
    st.markdown(badge_html, unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-header">📅 Schedule File</div>', unsafe_allow_html=True)
        sched_file = st.file_uploader("Schedule", type=["xlsx","xls"], key="sched", label_visibility="collapsed")
    with c2:
        st.markdown('<div class="section-header">👥 Staff Roster File</div>', unsafe_allow_html=True)
        staff_file = st.file_uploader("Staff", type=["xlsx","xls"], key="staff", label_visibility="collapsed")

    st.markdown("")
    if st.button("▶ Build Roster", key="btn_step1"):
        if not sched_file or not staff_file:
            st.warning("Upload both files first.", icon="⚠️"); st.stop()

        with st.spinner("Reading files…"):
            sched_df = pd.read_excel(sched_file, sheet_name=0)
            staff_df = pd.read_excel(staff_file, sheet_name=0)

        if "Name" not in sched_df.columns:
            st.error("Schedule must have a **Name** column."); st.stop()
        if "EmployeeID" not in sched_df.columns:
            st.error("Schedule must have an **EmployeeID** column."); st.stop()
        missing = {"Name", "CSLoginName", "EmployeeNumber"} - set(staff_df.columns)
        if missing:
            st.error(f"Staff file missing: {missing}"); st.stop()

        with st.spinner("Matching employees…"):
            lookup = build_staff_lookup(staff_df)
            records = []
            for _, row in sched_df.iterrows():
                matched, mtype = match_employee(row, lookup, staff_df)
                out = row.to_dict()
                for day in DAY_COLS:
                    if day in out:
                        out[day] = normalize_schedule(out[day])
                if matched is not None:
                    loc_raw = matched.get("On Site Location", "")
                    loc = loc_raw if (pd.notna(loc_raw) and str(loc_raw).strip()) else "WFH"
                    out.update({
                        "FullName_Staff": matched.get("Name", ""),
                        "CSLoginName": matched.get("CSLoginName", ""),
                        "EmployeeNumber": matched.get("EmployeeNumber", ""),
                        "OnSiteLocation": loc,
                        "MatchType": mtype
                    })
                else:
                    out.update({
                        "FullName_Staff": "", "CSLoginName": "", "EmployeeNumber": "",
                        "OnSiteLocation": "", "MatchType": mtype
                    })
                records.append(out)

        roster = pd.DataFrame(records)
        before = len(roster)
        roster = roster[~roster["OnSiteLocation"].str.strip().str.lower().isin(EXCLUDED_LOCATIONS)]
        st.session_state.roster = roster

        st.markdown("---")
        m1, m2, m3, m4 = st.columns(4)
        for col, val, lbl, cls in [
            (m1, len(roster), "Roster Size", ""),
            (m2, roster["MatchType"].str.startswith("exact").sum(), "Exact", "matched"),
            (m3, roster["MatchType"].str.startswith("fuzzy").sum(), "Fuzzy", "fuzzy"),
            (m4, roster["MatchType"].isin(["unmatched","no_name"]).sum(), "Unmatched", "unmatched"),
        ]:
            with col:
                st.markdown(f'<div class="metric-card"><div class="metric-val {cls}">{val}</div><div class="metric-lbl">{lbl}</div></div>', unsafe_allow_html=True)

        st.success("✔ Roster built! Switch to **Step 2**.")
        disp = [c for c in ["Name","FullName_Staff","CSLoginName","EmployeeNumber","OnSiteLocation","HireStatus","PositionName","MatchType"] + DAY_COLS if c in roster.columns]
        st.dataframe(clean_for_display(roster[disp]), width="stretch", height=380)

        d1, d2 = st.columns([2,1])
        with d1:
            st.download_button("⬇️ Download Roster as Excel", to_excel_bytes(roster), "roster.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with d2:
            st.download_button("⬇️ Download as CSV", roster.to_csv(index=False).encode(), "roster.csv", "text/csv", use_container_width=True)

    elif not roster_ready:
        st.markdown("""
        <div style="background:#1a1d27;border:1px solid #2a2d3a;border-radius:12px;padding:2rem;text-align:center;color:#7c7f8e;margin-top:1rem;">
            <div style="font-size:2.5rem;margin-bottom:0.5rem;">📋</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:1.1rem;color:#e8eaf0;">Upload Schedule + Staff to build the roster</div>
        </div>
        """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 — FAST vectorized processing with dict lookups
# ═══════════════════════════════════════════════════════════════════════════════

with tab2:
    st.markdown("")
    roster = st.session_state.roster

    if roster is None:
        st.markdown("""
        <div style="background:#1a1d27;border:1px solid #2a2d3a;border-radius:12px;padding:2rem;text-align:center;color:#7c7f8e;margin-top:1rem;">
            <div style="font-size:2.5rem;margin-bottom:0.5rem;">⏳</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:1.1rem;color:#e8eaf0;">Complete Step 1 first</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown('<span class="step-badge done">✔ Roster loaded</span>', unsafe_allow_html=True)
        st.markdown(f'<span style="color:#7c7f8e;font-size:0.82rem;">{len(roster):,} roster rows</span>', unsafe_allow_html=True)

        st.markdown("")
        st.markdown('<div class="section-header">⏱️ Timesheet File(s) *</div>', unsafe_allow_html=True)
        ts_files = st.file_uploader("Timesheets", type=["xlsx","xls"], accept_multiple_files=True, key="ts", label_visibility="collapsed")

        st.markdown("")
        st.markdown('<div class="section-header">🌴 Leave File(s)</div>', unsafe_allow_html=True)
        leave_files = st.file_uploader("Leave files", type=["xlsx","xls"], accept_multiple_files=True, key="leave", label_visibility="collapsed")

        st.markdown("")
        st.markdown('<div class="section-header">👤 Headcount File(s) *</div>', unsafe_allow_html=True)
        hc_files = st.file_uploader("Headcount", type=["xlsx","xls"], accept_multiple_files=True, key="hc_step2", label_visibility="collapsed")

        st.markdown("")
        st.markdown('<div class="section-header">📋 Override File (Optional)</div>', unsafe_allow_html=True)
        override_file = st.file_uploader("Override", type=["xlsx"], key="override_uploader", label_visibility="collapsed")

        tmpl1, tmpl2 = st.columns([1,1])
        with tmpl1:
            st.download_button("⬇️ Download Override Template", build_override_template(), "override_template.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with tmpl2:
            st.markdown('<div style="color:#7c7f8e;font-size:0.78rem;padding-top:0.4rem;">4 sheets: Schedule, Leave, Holidays, Exemptions</div>', unsafe_allow_html=True)

        st.markdown("")
        if st.button("▶ Build Final Attendance", key="btn_step2"):
            if not ts_files:
                st.warning("Upload at least one timesheet file.", icon="⚠️"); st.stop()
            if not hc_files:
                st.warning("Upload at least one headcount file.", icon="⚠️"); st.stop()

            # ─────────────────────────────────────────────────────────────────
            # 1. LOAD ALL FILES INTO MEMORY (dict-based)
            # ─────────────────────────────────────────────────────────────────
            with st.spinner("Loading files into memory…"):
                # Headcount files
                doj_lookup: dict[int, date] = {}
                hc_cache: dict[str, pd.DataFrame] = {}
                hc_dates: list[date] = []
                for f in hc_files:
                    df = pd.read_excel(f, sheet_name=0)
                    if "ECN" not in df.columns or "DOJ Knack" not in df.columns:
                        st.error(f"Headcount **{f.name}** missing ECN or DOJ Knack"); st.stop()
                    export_dt = extract_headcount_date(f.name, df)
                    dt_key = export_dt.strftime("%Y-%m-%d") if export_dt else f.name
                    df["ECN"] = pd.to_numeric(df["ECN"], errors="coerce").dropna().astype(int)
                    for _, row in df.iterrows():
                        doj = parse_doj_knack(row.get("DOJ Knack"))
                        if doj:
                            doj_lookup[int(row["ECN"])] = doj
                    hc_cache[dt_key] = df
                    hc_dates.append(export_dt)

                if hc_dates:
                    st.info(f"📅 Headcount snapshots: {', '.join(d.strftime('%m/%d/%Y') for d in sorted(hc_dates) if d)}")

                # Leave files — build dict of (normalized_name, date) → leave_row
                leave_lookup: dict[tuple[str, str], dict] = {}
                for lf in leave_files:
                    ldate = normalize_filename_date(lf.name)
                    ldf = read_leave_file(lf)
                    for _, lrow in ldf.iterrows():
                        name = ""
                        for col in ["Name", "name", "Employee Name", "EmployeeName", "Full Name", "FullName"]:
                            if col in lrow and pd.notna(lrow[col]):
                                name = str(lrow[col]).strip()
                                if name:
                                    break
                        if name:
                            clean_name = normalize(name)
                            leave_lookup[(clean_name, ldate)] = lrow.to_dict()

                # Timesheets — build dict of (date, email_lower) → row_dict
                ts_lookup: dict[tuple[str, str], dict] = {}
                ts_dates: list[str] = []
                for tsf in ts_files:
                    date_str = normalize_filename_date(tsf.name)
                    ts_dates.append(date_str)
                    tdf = read_timesheet(tsf)
                    for _, trow in tdf.iterrows():
                        email = str(trow.get("Agent Email", "")).strip().lower()
                        if email:
                            ts_lookup[(date_str, email)] = trow.to_dict()

            # ─────────────────────────────────────────────────────────────────
            # 2. BUILD BASE ATTENDANCE DATAFRAME (vectorized, no Python loops)
            # ─────────────────────────────────────────────────────────────────
            with st.spinner("Building attendance records…"):
                # Determine which columns exist in roster for the core selection
                core_cols = ["Name", "CSLoginName", "EmployeeNumber", "OnSiteLocation",
                             "HireStatus", "PositionName"]
                available_core = [c for c in core_cols if c in roster.columns]
                day_cols_available = [c for c in DAY_COLS if c in roster.columns]
                roster_core = roster[available_core + day_cols_available].copy()

                # Build date info dict
                date_info = {}
                for d in ts_dates:
                    dt = parse_date_str(d)
                    date_info[d] = {
                        "date_obj": dt.date() if dt else None,
                        "day_col": dt.strftime("%a") if dt else None,
                    }

                # Repeat roster for each date
                all_records = []
                for d in ts_dates:
                    df_day = roster_core.copy()
                    df_day["Date"] = d
                    df_day["_date_obj"] = date_info[d]["date_obj"]
                    df_day["_day_col"] = date_info[d]["day_col"]
                    all_records.append(df_day)

                base_df = pd.concat(all_records, ignore_index=True)

                # ── Vectorized shift extraction ──
                def get_shift_for_day(row):
                    dc = row["_day_col"]
                    if dc and dc in row:
                        return normalize_schedule(row[dc])
                    return "Rest Day"

                base_df["Shift"] = base_df.apply(get_shift_for_day, axis=1)
                base_df["Is Scheduled"] = (base_df["Shift"] != "Rest Day").astype(int)

                # ── Vectorized timesheet join via dict lookup ──
                base_df["_cs_lower"] = base_df["CSLoginName"].astype(str).str.strip().str.lower()
                base_df["_ts_key"] = base_df.apply(lambda r: (normalize_filename_date(r["Date"]), r["_cs_lower"]), axis=1)
                base_df["_ts_row"] = base_df["_ts_key"].map(ts_lookup)

                # Extract timesheet fields vectorized
                base_df["AgentEmail"] = base_df["_ts_row"].apply(lambda x: x.get("Agent Email", "") if isinstance(x, dict) else "")
                base_df["FirstLogin"] = base_df["_ts_row"].apply(lambda x: x.get("First Login") if isinstance(x, dict) else None)
                base_df["LastLogout"] = base_df["_ts_row"].apply(lambda x: x.get("Last Logout", "") if isinstance(x, dict) else "")
                base_df["Active"] = base_df["_ts_row"].apply(lambda x: x.get("Active") if isinstance(x, dict) else None)
                base_df["Break"] = base_df["_ts_row"].apply(lambda x: x.get("Break", "") if isinstance(x, dict) else "")

                # ── Vectorized attendance status ──
                def vec_attendance_status(row):
                    fl = row["FirstLogin"]
                    act = row["Active"]
                    has_login = pd.notna(fl) and str(fl).strip() not in ("", "NaT")
                    try:
                        av = float(act) if pd.notna(act) else 0.0
                    except:
                        av = 0.0
                    if has_login and av >= 0.03:
                        return "Present"
                    if has_login and av < 0.03:
                        return "For Review"
                    if not has_login and av > 0:
                        return "For Review"
                    return ""

                base_df["_ts_status"] = base_df.apply(vec_attendance_status, axis=1)

                # ── Vectorized leave join via dict lookup ──
                base_df["_name_clean"] = base_df["Name"].apply(normalize)
                base_df["_leave_key"] = base_df.apply(lambda r: (r["_name_clean"], normalize_filename_date(r["Date"])), axis=1)
                base_df["_leave_row"] = base_df["_leave_key"].map(leave_lookup)

                # Extract leave status vectorized
                def vec_leave_status(leave_row):
                    if not isinstance(leave_row, dict):
                        return None
                    for col in ["SL", "UPL"]:
                        val = leave_row.get(col, 0)
                        if pd.notna(val):
                            try:
                                if float(val) > 0:
                                    return "Absent"
                            except:
                                if str(val).strip().lower() in ("1", "yes", "true", "y"):
                                    return "Absent"
                    for col in ["VL", "ML", "BL", "SPL", "PL", "MWL", "BDL", "HLD", "OFST", "PTO"]:
                        val = leave_row.get(col, 0)
                        if pd.notna(val):
                            try:
                                if float(val) > 0:
                                    return "Leave"
                            except:
                                if str(val).strip().lower() in ("1", "yes", "true", "y"):
                                    return "Leave"
                    return None

                base_df["_leave_status"] = base_df["_leave_row"].apply(vec_leave_status)

                # ── Vectorized flag calculation ──
                # Present = 1 ONLY if Active >= 0.1 AND has First Login
                base_df["Present"] = (base_df["_ts_status"] == "Present").astype(int)
                base_df["For Review"] = (base_df["_ts_status"] == "For Review").astype(int)
                base_df["Leave"] = (base_df["_leave_status"] == "Leave").astype(int)
                base_df["Absent"] = 0

                # RULE: If Present = 1, Leave MUST be 0 (Present wins over Leave)
                # This only applies when Is Scheduled = 1
                scheduled_and_present = (base_df["Is Scheduled"] == 1) & (base_df["Present"] == 1)
                base_df.loc[scheduled_and_present, "Leave"] = 0

                # For Review on Rest Day → not absent, not scheduled
                fr_rest = (base_df["For Review"] == 1) & (base_df["Shift"] == "Rest Day")
                base_df.loc[fr_rest, "Absent"] = 0
                base_df.loc[fr_rest, "Is Scheduled"] = 0

                # For Review on working day → absent, scheduled
                fr_work = (base_df["For Review"] == 1) & (base_df["Shift"] != "Rest Day")
                base_df.loc[fr_work, "Absent"] = 1
                base_df.loc[fr_work, "Is Scheduled"] = 1

                # Leave → absent=0 (but only if not already fixed by Present rule above)
                leave_mask = (base_df["Leave"] == 1)
                base_df.loc[leave_mask, "Absent"] = 0

                # Neither present, for_review, nor leave → absent
                base_df.loc[(base_df["Present"] == 0) & (base_df["For Review"] == 0) & (base_df["Leave"] == 0), "Absent"] = 1

                # DOJ Knack check (vectorized)
                base_df["_emp_num"] = pd.to_numeric(base_df["EmployeeNumber"], errors="coerce")
                base_df["_doj"] = base_df["_emp_num"].map(doj_lookup)
                not_yet = (base_df["_doj"].notna()) & (base_df["_date_obj"].notna()) & (base_df["_doj"] > base_df["_date_obj"])
                base_df.loc[not_yet, "Shift"] = "Not yet hired"
                base_df.loc[not_yet, "Is Scheduled"] = 0
                base_df.loc[not_yet, "Present"] = 0
                base_df.loc[not_yet, "For Review"] = 0
                base_df.loc[not_yet, "Leave"] = 0
                base_df.loc[not_yet, "Absent"] = 0
                for dc in DAY_COLS:
                    mask = not_yet & (base_df["_day_col"] == dc)
                    if dc in base_df.columns:
                        base_df.loc[mask, dc] = "Not yet hired"

                # Late calculation (vectorized where possible)
                base_df["_shift_start"] = base_df["Shift"].apply(parse_shift_start)
                base_df["_first_login_dt"] = base_df["FirstLogin"].apply(safe_parse_datetime)
                base_df["Late"] = 0
                base_df["Late Minutes"] = 0.0

                late_mask = base_df["_shift_start"].notna() & base_df["_first_login_dt"].notna() & base_df["_date_obj"].notna() & ~not_yet
                for idx in base_df[late_mask].index:
                    row = base_df.loc[idx]
                    threshold = datetime.combine(row["_date_obj"], row["_shift_start"])
                    diff = (row["_first_login_dt"] - threshold).total_seconds() / 60.0
                    if diff > 0:
                        base_df.at[idx, "Late"] = 1
                        base_df.at[idx, "Late Minutes"] = round(diff, 2)

            # ─────────────────────────────────────────────────────────────────
            # 3. MERGE HEADCOUNT (one merge per timesheet date, vectorized)
            # ─────────────────────────────────────────────────────────────────
            with st.spinner("Merging headcount data…"):
                all_merged = []
                for d in ts_dates:
                    day_df = base_df[base_df["Date"] == d].copy()
                    date_obj = date_info[d]["date_obj"]

                    # Pick headcount
                    hc_df = None
                    if date_obj and hc_dates:
                        best_dt = max((dt for dt in hc_dates if dt and dt <= date_obj), default=min(hc_dates))
                        hc_df = hc_cache[best_dt.strftime("%Y-%m-%d")]
                    else:
                        hc_df = next(iter(hc_cache.values()))

                    # Merge
                    hc_df = hc_df.copy()
                    hc_df["ECN"] = pd.to_numeric(hc_df["ECN"], errors="coerce").dropna().astype(int)
                    day_df["_merge_key"] = pd.to_numeric(day_df["EmployeeNumber"], errors="coerce").dropna().astype(int)

                    hc_cols = ["ECN", "Employee", "Project", "Sub-Process", "Supervisor",
                               "Role", "Manager", "DOJ Knack", "Date of Separation",
                               "Billable/Buffer", "Active/Inactive"]
                    hc_subset = hc_df[[c for c in hc_cols if c in hc_df.columns]].copy()

                    merged = day_df.merge(hc_subset, left_on="_merge_key", right_on="ECN", how="left")
                    merged = merged.drop(columns=["_merge_key", "ECN"], errors="ignore")
                    merged["HeadcountMatch"] = merged["Employee"].notna().map({True: "✅ Matched", False: "❌ Unmatched"})

                    # Separation date (vectorized)
                    if "Date of Separation" in merged.columns:
                        has_sep = merged["Date of Separation"].apply(lambda x: pd.notna(x) and str(x).strip() != "" and str(x).strip().lower() != "nat")
                        merged.loc[has_sep, "Absent"] = 0
                        merged.loc[has_sep, "Is Scheduled"] = 0

                    # Maternity/Suspended (vectorized)
                    if "Role" in merged.columns:
                        is_mat_susp = merged["Role"].astype(str).str.lower().str.contains("maternity|suspended", case=False, na=False)
                        with_login = is_mat_susp & (merged["Present"] == 1)
                        merged.loc[with_login, ["Is Scheduled", "Present", "Absent", "Leave"]] = [1, 1, 0, 0]
                        if "Active/Inactive" in merged.columns:
                            merged.loc[with_login, "Active/Inactive"] = "Active"
                        no_login = is_mat_susp & (merged["Present"] == 0)
                        merged.loc[no_login, ["Is Scheduled", "Absent", "Leave"]] = [0, 0, 0]

                    all_merged.append(merged)

                final_merged = pd.concat(all_merged, ignore_index=True)

            # ─────────────────────────────────────────────────────────────────
            # 4. OVERRIDES (vectorized)
            # ─────────────────────────────────────────────────────────────────
            if override_file is not None:
                with st.spinner("Applying overrides…"):
                    xl = pd.ExcelFile(override_file)

                    # Pre-parse for fast matching
                    final_merged["_date_parsed"] = final_merged["Date"].apply(lambda d: parse_override_date(str(d)) if pd.notna(d) else None)
                    final_merged["_emp_str"] = final_merged["EmployeeNumber"].astype(str).str.strip()
                    final_merged["_proj_lc"] = final_merged["Project"].astype(str).str.strip().str.lower()
                    final_merged["_sub_lc"] = final_merged["Sub-Process"].astype(str).str.strip().str.lower()

                    # Schedule overrides
                    if "Schedule" in xl.sheet_names:
                        sov = pd.read_excel(override_file, sheet_name="Schedule")
                        if {"ECN", "Effective Date", "Effective Until"}.issubset(set(sov.columns)):
                            sov["ECN"] = pd.to_numeric(sov["ECN"], errors="coerce").dropna().astype(int).astype(str)
                            sov["_s"] = sov["Effective Date"].apply(parse_override_date)
                            sov["_e"] = sov["Effective Until"].apply(parse_override_date)
                            for _, ov in sov.iterrows():
                                if ov["_s"] is None or ov["_e"] is None:
                                    continue
                                mask = (final_merged["_emp_str"] == ov["ECN"]) & (final_merged["_date_parsed"].notna()) & (final_merged["_date_parsed"] >= ov["_s"]) & (final_merged["_date_parsed"] <= ov["_e"])
                                if not mask.any():
                                    continue
                                for dn in DAY_COLS:
                                    if dn in ov and dn in final_merged.columns:
                                        dm = mask & (final_merged["_date_parsed"].apply(lambda d: d.strftime("%a") if d else "") == dn)
                                        if dm.any():
                                            ns = normalize_schedule(ov[dn])
                                            final_merged.loc[dm, dn] = ns
                                            final_merged.loc[dm, "Shift"] = ns
                                            final_merged.loc[dm, "Is Scheduled"] = 0 if ns == "Rest Day" else 1

                    # Leave overrides
                    if "Leave" in xl.sheet_names:
                        lov = pd.read_excel(override_file, sheet_name="Leave")
                        if {"ECN", "Date", "Leave"}.issubset(set(lov.columns)):
                            lov["ECN"] = pd.to_numeric(lov["ECN"], errors="coerce").dropna().astype(int).astype(str)
                            lov["_d"] = lov["Date"].apply(parse_override_date)
                            for _, ov in lov.iterrows():
                                if ov["_d"] is None or int(float(ov.get("Leave", 0))) != 1:
                                    continue
                                mask = (final_merged["_emp_str"] == ov["ECN"]) & (final_merged["_date_parsed"] == ov["_d"])
                                if not mask.any():
                                    continue
                                final_merged.loc[mask, ["Leave", "Absent", "Present", "For Review"]] = [1, 0, 0, 0]
                                nh = mask & (final_merged["Shift"] == "Not yet hired")
                                if nh.any():
                                    dn = ov["_d"].strftime("%a")
                                    if dn in final_merged.columns:
                                        final_merged.loc[nh, dn] = "Leave"
                                    final_merged.loc[nh, "Shift"] = "Leave"

                    # Holiday overrides
                    if "Holidays" in xl.sheet_names:
                        hov = pd.read_excel(override_file, sheet_name="Holidays")
                        if {"Project", "Date"}.issubset(set(hov.columns)):
                            hov["_p"] = hov["Project"].astype(str).str.strip().str.lower()
                            hov["_s"] = hov["Sub-Process"].fillna("").astype(str).str.strip().str.lower()
                            hov["_d"] = hov["Date"].apply(parse_override_date)
                            for _, ov in hov.iterrows():
                                if ov["_d"] is None or not ov["_p"]:
                                    continue
                                mask = (final_merged["_proj_lc"] == ov["_p"]) & (final_merged["_date_parsed"] == ov["_d"])
                                if ov["_s"]:
                                    mask = mask & (final_merged["_sub_lc"] == ov["_s"])
                                if not mask.any():
                                    continue
                                final_merged.loc[mask, ["Is Scheduled", "Absent", "Leave"]] = [0, 0, 0]
                                pm = mask & (final_merged["Present"] == 1)
                                final_merged.loc[pm, "Is Scheduled"] = 1
                                sm = mask & (~final_merged["Shift"].isin(["Not yet hired", "Leave"]))
                                final_merged.loc[sm, "Shift"] = "Holiday"
                                dn = ov["_d"].strftime("%a")
                                if dn in final_merged.columns:
                                    final_merged.loc[sm, dn] = "Holiday"

                    # Exemptions
                    if "Exemptions" in xl.sheet_names:
                        eov = pd.read_excel(override_file, sheet_name="Exemptions")
                        if {"ECN", "Exemption Date", "Effective Until", "Active/Inactive"}.issubset(set(eov.columns)):
                            eov["ECN"] = pd.to_numeric(eov["ECN"], errors="coerce").dropna().astype(int).astype(str)
                            eov["_s"] = eov["Exemption Date"].apply(parse_override_date)
                            eov["_e"] = eov["Effective Until"].apply(parse_override_date)
                            for _, ov in eov.iterrows():
                                if ov["_s"] is None or ov["_e"] is None:
                                    continue
                                mask = (final_merged["_emp_str"] == ov["ECN"]) & (final_merged["_date_parsed"].notna()) & (final_merged["_date_parsed"] >= ov["_s"]) & (final_merged["_date_parsed"] <= ov["_e"])
                                if not mask.any():
                                    continue
                                final_merged.loc[mask, ["Is Scheduled", "Absent"]] = [0, 0]
                                ns = str(ov.get("Active/Inactive", "")).strip()
                                if "Active/Inactive" in final_merged.columns and ns:
                                    final_merged.loc[mask, "Active/Inactive"] = ns

                    final_merged = final_merged.drop(columns=["_date_parsed", "_emp_str", "_proj_lc", "_sub_lc"], errors="ignore")
                st.success("✔ Overrides applied!")

            # ── Final consistency ──
            # Priority: Present > Leave > For Review > Absent
            # If Scheduled = 1 and Present = 1, clear Leave
            sched_present = (final_merged["Is Scheduled"] == 1) & (final_merged["Present"] == 1)
            final_merged.loc[sched_present, "Leave"] = 0

            um = final_merged["HeadcountMatch"] == "❌ Unmatched"
            final_merged.loc[um, ["Is Scheduled", "Absent"]] = [0, 0]

            ns = final_merged["Is Scheduled"] == 0
            final_merged.loc[ns, ["Absent", "Leave"]] = [0, 0]

            pns = (final_merged["Is Scheduled"] == 0) & (final_merged["Present"] == 1)
            final_merged.loc[pns, "Is Scheduled"] = 1

            frm = (final_merged["For Review"] == 1) & (final_merged["Is Scheduled"] == 1)
            final_merged.loc[frm, "Absent"] = 1

            lm = final_merged["Leave"] == 1
            final_merged.loc[lm, "Absent"] = 0

            # Clean up temp columns
            drop_cols = [c for c in ["_cs_lower", "_ts_key", "_ts_row", "_name_clean", "_leave_key", "_leave_row",
                         "_emp_num", "_doj", "_shift_start", "_first_login_dt", "_date_obj", "_day_col", "_ts_status", "_leave_status"] if c in final_merged.columns]
            final_merged = final_merged.drop(columns=drop_cols, errors="ignore")

            st.session_state.attendance = final_merged
            st.session_state.merged_headcount = final_merged

            # ── Metrics ──
            st.markdown("---")
            m1, m2, m3, m4, m5, m6, m7 = st.columns(7)
            for col, val, lbl, cls in [
                (m1, len(final_merged), "Total Rows", ""),
                (m2, final_merged["Present"].sum(), "Present", "matched"),
                (m3, final_merged["For Review"].sum(), "For Review", "review"),
                (m4, final_merged["Leave"].sum(), "On Leave", "neutral"),
                (m5, final_merged["Absent"].sum(), "Absent", "unmatched"),
                (m6, (final_merged["Shift"] == "Not yet hired").sum(), "Not Yet Hired", "not-hired"),
                (m7, (final_merged["HeadcountMatch"] == "✅ Matched").sum(), "HC Matched", "matched"),
            ]:
                with col:
                    st.markdown(f'<div class="metric-card"><div class="metric-val {cls}">{val}</div><div class="metric-lbl">{lbl}</div></div>', unsafe_allow_html=True)

            if (final_merged["HeadcountMatch"] == "❌ Unmatched").sum() > 0:
                st.markdown(f'<div style="text-align:center;color:#f87171;font-size:0.85rem;">⚠️ {(final_merged["HeadcountMatch"] == "❌ Unmatched").sum():,} rows unmatched</div>', unsafe_allow_html=True)

            # ── Display ──
            disp = [c for c in (["Date","Name","CSLoginName","EmployeeNumber","OnSiteLocation",
                     "Day","Shift","Is Scheduled","Present","Absent","Leave","For Review",
                     "FirstLogin","LastLogout","Active","Break",
                     "Late","Late Minutes","HireStatus","PositionName",
                     "Employee", "Project", "Sub-Process", "Supervisor",
                     "Role", "Manager", "DOJ Knack", "Date of Separation",
                     "Billable/Buffer", "Active/Inactive",
                     "HeadcountMatch"] + DAY_COLS) if c in final_merged.columns]

            tabs = st.tabs(["All", "✅ Present", "🟠 For Review", "🌴 Leave", "❌ Absent", "🔵 Not Yet Hired", "✅ HC Matched", "❌ HC Unmatched"])
            filters = [
                (tabs[0], None),
                (tabs[1], final_merged["Present"] == 1),
                (tabs[2], final_merged["For Review"] == 1),
                (tabs[3], final_merged["Leave"] == 1),
                (tabs[4], final_merged["Absent"] == 1),
                (tabs[5], final_merged["Shift"] == "Not yet hired"),
                (tabs[6], final_merged["HeadcountMatch"] == "✅ Matched"),
                (tabs[7], final_merged["HeadcountMatch"] == "❌ Unmatched"),
            ]
            for tab, filt in filters:
                with tab:
                    df_show = final_merged[filt] if filt is not None else final_merged
                    st.dataframe(clean_for_display(df_show[disp]), width="stretch", height=420)

            # ── Download ──
            st.markdown("---")
            d1, d2 = st.columns([2,1])
            with d1:
                out_name = ts_files[0].name.replace(".xlsx","").replace(".xls","") + ".xlsx"
                st.download_button("⬇️ Download Final as Excel", to_excel_bytes(final_merged), out_name,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with d2:
                st.download_button("⬇️ Download as CSV", final_merged.to_csv(index=False).encode(),
                    out_name.replace(".xlsx",".csv"), "text/csv", use_container_width=True)
